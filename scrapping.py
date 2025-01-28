# Cores
main = "\033[0;37;40"
red = "\033[0;31;40m"
green = "\033[0;32;40m"
yellow = "\033[0;33;40m"
blue = "\033[0;34;40m"
u_red = "\033[0;37;41m"
u_green = "\033[0;37;42m"
# Obtém os dados do site e salva
def get_data(arq_local="arq.xlsx"):
    # Importação de bilbiotecas
    from selenium import webdriver # Driver (controlador do navegador)
    from selenium.webdriver.common.by import By # Utilizado para indicar as referências internas
    from selenium.webdriver.common.keys import Keys # Algumas referências
    from selenium.webdriver.chrome.options import Options # Opções de configurações para o webdriver
    from selenium.webdriver.support.ui import WebDriverWait # modulo para fazer o código esperar enquanto realiza algum processo
    from selenium.webdriver.support import expected_conditions as EC # Condições pré-definidas para a interação
    import pandas as pd # Biblioteca para interação com as planilhas excel
    from time import sleep # Esperar

    # Scrapping part

    # Declaração de variáveis globais

    lista_cia = [] # Lista os ticker das ações 
    lista_tipo = [] # Lista os tipos de ações (BDR, FII, stock)
    relatorio = [] # Lista com listas para carregar no excel
    resumo = [] # Variável para guardar informações temporariamente
    titulos = ["Ticker", "Cotação"] # Lista de referencia para os títulos na tabela
    fii_titulos = ["Ticker"] # Cabeçalho próprio para os FIIs
    relatorio_fii = [] # Lista própria para os FIIs
    count = 0 # Contador ao longo do códgio
    conta_fii = 0
    conta_acao = 0
    deu_certo = False

    # Abrir o excel e obter tabelas internas para manipulação
    arquivo = pd.read_excel(arq_local, sheet_name=None)
    arq = arquivo["Referências"] 
    # Inicializar driver
    chrome_options = Options()
    nav = webdriver.Chrome(options=chrome_options)

    # Declaração de funções pontuais
    def get_essential_info(fim_cont=34):
        WebDriverWait(nav, 15).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "w-50")))    
        blocos = nav.find_elements(By.CLASS_NAME, "w-50")
        print(f"{main} Blocos encontrados: {red}{len(blocos)}")  # Depuração: Verificar quantos blocos foram encontrados
        print(f"{blue} Contando até o bloco número {yellow}{fim_cont}")
        resumo.append(str(lista_cia[count]).upper())  # Adiciona o ticker ao início do resumo
        WebDriverWait(nav, 15).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "special")))
        for a, b in enumerate(nav.find_elements(By.CLASS_NAME, "special")):
            cotate = b.find_element(By.TAG_NAME, "strong")
            resumo.append(cotate.text)
        contar = 0
        for i, item in enumerate(blocos):
            try:
                print(f"{yellow}-----------------------------------------")
                print(f"{main} Processando bloco {green}{i + 1}/{len(blocos)}")  # Depuração
                # Obtém o título
                if conta_acao == 0:  # Apenas na primeira iteração, armazena títulos
                    title = item.find_element(By.TAG_NAME, "h3").text  # Obtém o texto do elemento
                    print(f"{main} Título encontrado: {blue}{title}")  # Depuração
                    titulos.append(title)  # Adiciona ao cabeçalho

                # Obtém o conteúdo
                content = item.find_element(By.TAG_NAME, "strong").text  # Obtém o texto do elemento
                print(f"{green} Conteúdo encontrado: {content}")  # Depuração
                resumo.append(content)  # Adiciona ao resumo
                if i >= fim_cont:
                    break
            except Exception as e:
                contar += 1
                print(f"{red} Erro ao processar bloco {i + 1}: {e}")  # Captura qualquer exceção
        if contar <= 5:
            print(f"{yellow} -{titulos}-")
            return True
        else:
            return False

    def fii_get_essential_info(fim_cont=11):
        WebDriverWait(nav, 15).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "info")))    
        blocos = nav.find_elements(By.CLASS_NAME, "info")
        print(f"{main} Blocos encontrados: {red}{len(blocos)}")  # Depuração: Verificar quantos blocos foram encontrados
        print(f"{blue} Contando até o bloco número {yellow}{fim_cont}")
        resumo.append(str(lista_cia[count]).upper())  # Adiciona o ticker ao início do resumo
        contar = 0
        for i, item in enumerate(blocos):
            try:
                print(f"{yellow}-----------------------------------------")
                print(f"{main} Processando bloco {green}{i + 1}/{len(blocos)}")  # Depuração
                # Obtém o título
                if conta_fii == 0:  # Apenas na primeira iteração, armazena títulos
                    title = item.find_element(By.TAG_NAME, "h3").text  # Obtém o texto do elemento
                    print(f"{main} Título encontrado: {blue}{title}")  # Depuração
                    fii_titulos.append(title)  # Adiciona ao cabeçalho

                # Obtém o conteúdo
                content = item.find_element(By.TAG_NAME, "strong").text  # Obtém o texto do elemento
                print(f"{green} Conteúdo encontrado: {content}")  # Depuração
                resumo.append(content)  # Adiciona ao resumo
                if i >= fim_cont:
                    break
            except Exception as e:
                contar += 1
                print(f"{red} Erro ao processar bloco {i + 1}: {e}")  # Captura qualquer exceção
        if contar <= 5:
            return True
        else:
            return False

    # Obtenção dos nomes das empresas a pesquisar
    for _, row in arq.iterrows(): 
        lista_cia.append(str(row["ticker"]).lower())
        lista_tipo.append(str(row["tipo"]).lower())
    if len(lista_cia) >= 2:
        print(f"{u_green} Dados lidos com sucesso")
    print(f"{u_green} Prosseguindo para obtenção na Web")

    # Obtenção das informações na web
    # Leitura e localização
    while True: 
        # Esquematizar um try para localizar possíveis erros
        try:
            # Abre o site em questão
            nav.get("https://statusinvest.com.br")
            # Espera para inicialização
            sleep(5)
            # Fechamento do pop up
            try:
                # Esperar o pop-up aparecer
                WebDriverWait(nav, 5).until(EC.element_to_be_clickable((By.CLASS_NAME, "btn-close")))
                # Fechar o pop-up
                close_button = nav.find_element(By.CLASS_NAME, "btn-close")
                close_button.click()
                print(f"{blue} Pop-up fechado com sucesso.")
            except Exception as e:
                print(f"{u_red} Erro ao fechar o pop-up: {e}")
            # Abrir a barra de pesquisa
            nav.find_element(By.CLASS_NAME, "main-search").click()
            # Aguarde a barra de pesquisa estar disponível
            WebDriverWait(nav, 13).until(EC.presence_of_element_located((By.CLASS_NAME, "input-form")))
            WebDriverWait(nav, 13).until(EC.presence_of_element_located((By.CLASS_NAME, "twitter-typeahead")))
            # Inserir o texto correspondente
            try:
                # Localiza o campo de texto
                search_input = nav.find_element(By.CSS_SELECTOR, ".Typeahead-input.input.tt-input")
                # Insere o texto e confirma
                search_input.send_keys(lista_cia[count])
                search_input.send_keys(Keys.RETURN)
            except:
                print(f"{red} Não foi possível preencher o campo")
            else:
                print(f"{green} Campo preenchido")

            # Identifica o tipo de item e acessa seu link correspondente    
            try:
                if lista_tipo[count] == "ação": # Script para acessar uma ação por pesquisa
                    link = WebDriverWait(nav, 5).until(EC.element_to_be_clickable((By.XPATH, f"//a[contains(@href, '/acoes/{lista_cia[count]}')]")))
                    nav.execute_script("arguments[0].click();", link)
                    nav.find_element(By.XPATH, f"//a[contains(@href, '/acoes/{lista_cia[count]}')]").click()
                elif lista_tipo[count] == "fii": # Script para acessar um FII por pesquisa
                    link = WebDriverWait(nav, 5).until(EC.element_to_be_clickable((By.XPATH, f"//a[contains(@href, '/fundos-imobiliarios/{lista_cia[count]}')]")))
                    nav.execute_script("arguments[0].click();", link)
                    nav.find_element(By.XPATH, f"//a[contains(@href, '/fundos-imobiliarios/{lista_cia[count]}')]").click()
                elif lista_tipo[count] == "bdr": # Script para acessa uma stock por pesquisa
                    link = WebDriverWait(nav, 5).until(EC.element_to_be_clickable((By.XPATH, f"//a[contains(@href, '/acoes/eua/{lista_cia[count]}')]")))
                    nav.execute_script("arguments[0].click();", link)
                    nav.find_element(By.XPATH, f"//a[contains(@href, '/acoes/eua/{lista_cia[count]}')]").click()
            except:
                print(f"{u_red} Não foi possível acessar o ativo {lista_cia[count]}")
    
            # Listagem das informações para o Excel
            if lista_tipo[count] == "fii":  # Listagem para FIIs
                deu_certo = fii_get_essential_info()
            else:  # Listagem para ações e Stocks
                deu_certo = get_essential_info()
            if deu_certo == True:
                print(f"{blue} {str(lista_cia[count]).upper()} registrado com sucesso.")
            if deu_certo == False:
                print(f"{red} {str(lista_cia[count]).upper()} não foi registrado completamente.")

        except:
            print(f"{u_red} Erro fatal no ativo de ticker {str(lista_cia[count]).upper()}") 
        finally:
            # Adicionar uma lista para mandar pro excel depois
            if resumo != []:
                if lista_tipo[count] == "fii":
                    if len(fii_titulos) >= 3:
                        relatorio_fii.append(fii_titulos.copy())
                    relatorio_fii.append(resumo.copy())
                    conta_fii += 1
                else:     
                    if len(titulos) >= 3:
                        relatorio.append(titulos.copy())
                    relatorio.append(resumo.copy())
                    conta_acao += 1
                resumo.clear()
                titulos = ["Ticker", "Cotação"]
                fii_titulos = ["Ticker"]
                count += 1
                print(f"{blue}-------------------------------{main}")
                for a in range(len(relatorio)):
                    print(relatorio[a])
                print(f"{blue}-------------------------------{main}")
                for a in range(len(relatorio_fii)):
                    print(relatorio_fii[a])
                print(f"{yellow} Informação salva com sucesso!")
        # Repete
            if count >= len(lista_cia):
                print(f"{u_green} Coleta de dados concluída\n\n")
                print(f"{main} ...")
                sleep(1)
                break
            else:
                nav = webdriver.Chrome(options=chrome_options)

    # Formatação das informações
    # Inserir informações na planilha excel
    nav.quit()
    print(f"{main} Transferindo para a panilha {arq_local}")
    # Problema crônico na coleta de dados de FIIs
    # Correção temporária
    for correct in range(len(relatorio_fii)):
        if relatorio_fii[correct] == []:
            relatorio_fii[correct] = [f"{lista_cia[correct]}", "erro fatal"]
    refer = pd.DataFrame(relatorio[1:], columns=relatorio[0])
    with pd.ExcelWriter(arq_local, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        refer.to_excel(writer, sheet_name="Ações", index=False)
    print(f"{green} Dados de ações salvos com sucesso na planilha excel!")
    refer2 = pd.DataFrame(relatorio_fii[1:], columns=relatorio_fii[0])
    with pd.ExcelWriter(arq_local, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        refer2.to_excel(writer, sheet_name="FIIs", index=False)
    print(f"{green} Dados de FIIs salvos com sucesso na planilha excel!")
# Calcula as notas para organizar a planilha
def calculate_nota(value, all_values, higher_is_better=True):
    """
    Calculate a score for a single value based on a list of all values and the given criterion.

    Args:
        value (float): The specific value to calculate the score for.
        all_values (list of float): The complete list of values to determine scaling.
        higher_is_better (bool): Whether higher values are better. If False, lower values are better.

    Returns:
        float: The calculated score mapped to a scale of 0 to 10.
    """
    import numpy as np

    # Calculate mean and standard deviation from all_values
    mean = np.mean(all_values)
    std_dev = np.std(all_values)

    # Handle edge case where all values are the same
    if std_dev == 0:
        return 5.0  # Neutral score if no variation

    # Calculate Z-scores for all values
    z_scores = [(v - mean) / std_dev for v in all_values]

    # Adjust for "lower is better" criterion
    if not higher_is_better:
        z_scores = [-z for z in z_scores]

    # Determine z_min and z_max
    z_min = min(z_scores)
    z_max = max(z_scores)

    # Calculate the Z-score for the specific value
    z_value = (value - mean) / std_dev
    if not higher_is_better:
        z_value = -z_value

    # Map the Z-score to the scale of 0 to 10
    score = 10 * (z_value - z_min) / (z_max - z_min)

    return score
# Organiza os dados dentro da planilha em tabelas específicas
def manage_data(selic, arq_local="arq.xlsx"):
    import pandas as pd
    from stocks import ativo
    from math import sqrt
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
    # Declaração de variáveis do código
    acao = []
    fii = []
    # Abrir o excel e obter tabelas internas para manipulação
    arquivo = pd.read_excel(arq_local, sheet_name=None)
    arq = arquivo["Referências"] 
    local_acao = arquivo["Ações"]
    local_fii = arquivo["FIIs"]

    for i, row in local_acao.iterrows(): 
        if not row.isnull().all():
            linha = [str(valor).upper() if pd.notnull(valor) else "" for valor in row]  # Converte NaN para string vazia
            act = ativo()
            act.__init__()
            act.cria_obj(1, linha, selic=selic)
            acao.append(act)
    for i, row in local_fii.iterrows(): 
        if not row.isnull().all():
            linha = [str(valor).upper() if pd.notnull(valor) else "" for valor in row]  # Converte NaN para string vazia
            act = ativo()
            act.__init__()
            act.cria_obj(2, linha, selic=selic)
            fii.append(act)
    # Totais de valores FIIs
    total_divy = []
    total_dycagr = []
    total_valorcagr = []
    total_dif = []
    total_pvp = [] # P/VP
    for um in range(len(fii)):
        total_divy.append(fii[um].dy) # Valor DY
        total_dycagr.append(fii[um].dycagr)
        total_valorcagr.append(fii[um].valorcagr)
        total_dif.append(fii[um].patrimonio-fii[um].cotacao)
        total_pvp.append(fii[um].pvp)
    # Organiza FIIs
    multi = 1
    for dois in range(len(total_dif)):
        if fii[dois].classe == 1:
            multi = 1
        if fii[dois].classe == 2:
            multi = 2.5
        divy = multi * (calculate_nota(fii[dois].dy, total_divy, True))
        pvp = calculate_nota(fii[dois].pvp, total_pvp, False)
        dycagr = calculate_nota(fii[dois].dycagr, total_dycagr, True)
        valorcagr = calculate_nota(fii[dois].valorcagr, total_valorcagr, True)
        dif = calculate_nota(fii[dois].patrimonio-fii[dois].cotacao, total_dif, True)
        media = (divy+pvp+dycagr+valorcagr+dif)/(4+multi)
        fii[dois].nota_geral.append(media)    
    # Totais de valores ações
    total_pvp.clear()
    total_valoriza = [] # Valorização com base na fórmula de graham (a-b)
    total_dy = [] # Dividend yeld
    total_roe = [] # ROE
    total_roic = [] # ROIC
    total_pl = [] # P/L
    total_pegr = [] # PEG Ratio
    total_divebitda = [] # Div líquida/EBITDA
    total_divpl = [] # Div líquida/PL
    total_lc = [] # Líquidez corrente
    total_pa = [] # Passivos/Ativos
    total_relation = [] # Relação pronta
    for item in range(len(acao)):
    #Eficiência (2)
        # DY (deve ser menor que o ROA)
        total_dy.append(acao[item].dyroa)
        # ROE (comparável com a Selic)    
        total_roe.append(acao[item].roe)
        #	ROIC (quanto maior melhor)
        total_roic.append(acao[item].roic)
    #Valuation (1)
        # P/L (quanto menor melhor, mas acima de 0)
        total_pl.append(acao[item].pl)
        # PEG Ratio (quanto menor melhor, mas acima de 0)
        total_pegr.append(acao[item].pegr)
        # P/VP (quanto menor melhor, mas acima de 0)
        total_pvp.append(acao[item].pvp)
        # Fórmula de Benjamin Graham - cotação (quanto maior melhor)
        total_valoriza.append(acao[item].graham-acao[item].cotacao)
    #Endividamento (4)
        # Dívida líquida/EBITDA (ideal abaixo de 3)
        total_divebitda.append(acao[item].divebitda)
        # Dívida Líquida/PL (alavancagem: quanto menor, melhor)
        total_divpl.append(acao[item].divpl)
        # Líquidez corrente (quanto maior melhor)
        total_lc.append(acao[item].lc)
        # Passivos/ Ativos (quanto menor melhor)
        total_pa.append(acao[item].pa)
    #Crescimento (3)
        # Cria relação CAGR lucros/CAGR receitas (quanto mais próximo de 1, melhor)
        total_relation.append(acao[item].relation)
    # Organiza notas e separa
    eficiencia = 0
    valuation = 0
    divida = 0
    cresce = 0
    for outro in range(len(total_relation)):
        if acao[outro].classe == 1:
        #Eficiência (2)
            dy = calculate_nota(acao[outro].dyroa, total_dy, True) # DY (quanto maior melhor)
            roe = calculate_nota(acao[outro].roe, total_roe, True) # ROE (quanto maior melhor)
            roic = calculate_nota(acao[outro].roic, total_roic, True) # ROIC (quanto maior melhor)
            eficiencia = (dy+roe+roic)/3
            acao[outro].nota_geral.append(eficiencia)
        #Valuation (1)
            pl = calculate_nota(acao[outro].pl, total_pl, False) # P/L (quanto menor melhor, mas acima de 0)
            pegr = calculate_nota(acao[outro].pegr, total_pegr, False) # PEG Ratio (quanto menor melhor, mas acima de 0)
            pvp = calculate_nota(acao[outro].pvp, total_pvp, False) # P/VP (quanto menor melhor, mas acima de 0)
            valoriza = calculate_nota(acao[outro].graham-acao[outro].cotacao, total_valoriza, True) # Fórmula de Benjamin Graham - cotação (quanto maior melhor)
            valuation = (pl+pegr+pvp+valoriza)/4
            acao[outro].nota_geral.append(valuation)
        #Endividamento (4)
            divebitda = calculate_nota(acao[outro].divebitda, total_divebitda, False) # Dívida líquida/EBITDA (ideal abaixo de 3)
            divpl = calculate_nota(acao[outro].divpl, total_divpl, False) # Dívida Líquida/PL (alavancagem: quanto menor, melhor)
            lc = calculate_nota(acao[outro].lc, total_lc, True) # Líquidez corrente (quanto maior melhor)
            pa = calculate_nota(acao[outro].pa, total_pa, False) # Passivos/ Ativos (quanto menor melhor)
            divida = (divebitda+divpl+lc+pa)/4
            acao[outro].nota_geral.append(divida)
        #Crescimento (3)
            # Cria relação CAGR lucros/CAGR receitas (quanto mais próximo de 1, melhor) 
            cresce = calculate_nota(acao[outro].relation, total_relation, False)
            acao[outro].nota_geral.append(cresce)
            media_geral = (cresce + divida + valuation +eficiencia)/4
            acao[outro].nota_geral.append(media_geral)
        if acao[outro].classe == 2:
            #Tecnologia e Small caps
        #Eficiência (1)
            dy = calculate_nota(acao[outro].dyroa, total_dy, True) # DY (quanto maior melhor)
            roe = calculate_nota(acao[outro].roe, total_roe, True) # ROE (quanto maior melhor)
            roic = calculate_nota(acao[outro].roic, total_roic, True) # ROIC (quanto maior melhor)
            eficiencia = (dy+roe+roic)/3
            acao[outro].nota_geral.append(eficiencia)
        #Valuation (3)
            pl = calculate_nota(acao[outro].pl, total_pl, False) # P/L (quanto menor melhor, mas acima de 0)
            pegr = calculate_nota(acao[outro].pegr, total_pegr, False) # PEG Ratio (quanto menor melhor, mas acima de 0)
            valuation = (pl+pegr)/2
            acao[outro].nota_geral.append(valuation)
        #Endividamento (2)
            divebitda = calculate_nota(acao[outro].divebitda, total_divebitda, False) # Dívida líquida/EBITDA (ideal abaixo de 3)
            divpl = calculate_nota(acao[outro].divpl, total_divpl, False) # Dívida Líquida/PL (alavancagem: quanto menor, melhor)
            lc = calculate_nota(acao[outro].lc, total_lc, True) # Líquidez corrente (quanto maior melhor)
            pa = calculate_nota(acao[outro].pa, total_pa, False) # Passivos/ Ativos (quanto menor melhor)
            divida = (divebitda+divpl+lc+pa)/4
            acao[outro].nota_geral.append(divida)
        #Crescimento (4)
            # Cria relação CAGR lucros/CAGR receitas (quanto mais próximo de 1, melhor) 
            cresce = calculate_nota(acao[outro].relation, total_relation, False)
            acao[outro].nota_geral.append(cresce)
            media_geral = (cresce + divida + valuation +eficiencia)/4
            acao[outro].nota_geral.append(media_geral)
    # Adicionar na planilha o ticker de cada item, as notas e classificar
    new_acao = [] # Ticker, cotação, notas (5)
    new_fii = [] # Ticker, cotação, nota
    for icon in acao:
        calc = 22.5*icon.vpa*icon.lpa
        graham = sqrt(calc)
        new_acao.append({"Ticker":icon.nome, "Cotação":icon.cotacao, "preco alvo":graham, "nota_geral":icon.nota_geral, "margem":(graham-icon.cotacao), "margem%":((graham - icon.cotacao)/icon.cotacao)})
    for ite in fii:
        new_fii.append({"Ticker":ite.nome, "Cotação":ite.cotacao, "preco alvo":ite.patrimonio, "nota_geral":ite.nota_geral, "margem":(ite.patrimonio - ite.cotacao), "margem%":((ite.patrimonio-ite.cotacao)/ite.cotacao)})
    # Função para ordenar a lista de objetos
    def classificar_ativos(ativos, ordem_criterios=[4,3,2,1,0], reverso=True):
        # Função de chave para ordenação
        def chave_ordenacao(ativo):
            notas = ativo["nota_geral"]
            return tuple(-notas[i] for i in ordem_criterios)
        # Ordena os ativos com base nos critérios definidos
        return sorted(ativos, key=chave_ordenacao, reverse=reverso)
    ordem = [4, 3, 2, 1, 0]
    order = [0]
    # Ordena os ativos
    stock_order = classificar_ativos(new_acao, ordem, False)
    fii_order = classificar_ativos(new_fii, order, False)
    dados_fii = []
    dados_acao = []
    for ativos in stock_order:
        dados_acao.append({"Ticker": ativos["Ticker"],
            "Cotação":ativos["Cotação"],
            "Preço Alvo":ativos["preco alvo"],
            "Valuation": ativos["nota_geral"][0],
            "Eficiência": ativos["nota_geral"][1],
            "Crescimento": ativos["nota_geral"][2],
            "Endividamento": ativos["nota_geral"][3],
            "Nota final": ativos["nota_geral"][4],
            "Margem $":ativos["margem"], 
            "Margem %":ativos["margem%"]})
    for active in fii_order:
        dados_fii.append({"Ticker": active["Ticker"],
            "Preço Alvo":active["preco alvo"],
            "Cotação":active["Cotação"],
            "Nota final": active["nota_geral"][0],
            "Margem $":active["margem"],
            "Margem %":active["margem%"]})
    # Criar DataFrame a partir dos dados
    quadro_fii = pd.DataFrame(dados_fii)
    quadro_acao = pd.DataFrame(dados_acao)
    # Salvar o DataFrame em um arquivo Excel
    with pd.ExcelWriter(arq_local, mode="a", engine='openpyxl', if_sheet_exists="replace") as writer:
        quadro_acao.to_excel(writer, sheet_name='Make_stock', index=False)
        quadro_fii.to_excel(writer, sheet_name='Make_fii', index=False)
    make = load_workbook(arq_local) # Abre o arquivo pra fomatação
    # Define formatações para cada coluna, incluindo casas decimais
    abas_formatos = {
        "Make_stock": {
            "B": {"format": FORMAT_CURRENCY_USD_SIMPLE, "decimals": 2},  # Moeda com 2 casas decimais
            "C": {"format": FORMAT_CURRENCY_USD_SIMPLE, "decimals": 2},  # Moeda com 2 casas decimais
            "D": {"format": "0.00", "decimals": 2},  # Número com 2 casas decimais
            "E": {"format": "0.00", "decimals": 2},  # Número com 3 casas decimais
            "F": {"format": "0.00", "decimals": 2},  # Número com 2 casas decimais
            "G": {"format": "0.00", "decimals": 2},  # Número com 4 casas decimais
            "H": {"format": "0.00", "decimals": 2},  # Número inteiro
            "I": {"format": FORMAT_CURRENCY_USD_SIMPLE, "decimals": 2},  # Moeda com 2 casas decimais
            "J": {"format": FORMAT_PERCENTAGE, "decimals": 2},  # Porcentagem com 2 casas decimais
        },
        "Make_fii": {
            "B": {"format": FORMAT_CURRENCY_USD_SIMPLE, "decimals": 2},  # Moeda com 2 casas decimais
            "C": {"format": FORMAT_CURRENCY_USD_SIMPLE, "decimals": 2},  # Moeda com 2 casas decimais
            "D": {"format": "0.00", "decimals": 2},  # Número com 2 casas decimais
            "E": {"format": FORMAT_CURRENCY_USD_SIMPLE, "decimals": 2},  # Moeda com 2 casas decimais
            "F": {"format": FORMAT_PERCENTAGE, "decimals": 2},  # Porcentagem com 2 casas decimais
        },
    }

    # Aplica formatações automáticas
    for aba, colunas_formatos in abas_formatos.items():
        ws = make[aba]  # Selecionar aba pelo nome
        for col, formato_info in colunas_formatos.items():
            formato = formato_info["format"]
            decimals = formato_info["decimals"]
            for cell in ws[col][1:]:  # Ignora o cabeçalho
                if isinstance(cell.value, (int, float)):  # Apenas aplica em valores numéricos
                    cell.value = round(cell.value, decimals)  # Arredonda conforme as casas decimais
                cell.number_format = formato
                cell.alignment = Alignment(horizontal="center")

    # Salva o arquivo formatado
    make.save(arq_local)
# Abre a planilha conforme especificações
def abre_plan(arq_local="arq.xlsx"):
    import os
    print(f"{yellow} Abrindo a planilha excel.")
    os.startfile(arq_local)